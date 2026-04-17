"""
Kardz 商户后台订单监控脚本
- 自动登录，Session 过期自动重登
- 每 5 分钟轮询全部订单
- 新订单自动追加写入 Excel（手机号 + 邮箱 + 商品 + 金额 + 时间）
"""

import requests
import time
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============ 配置 ============
BASE_URL     = "https://ch-admin.kardz.com"
USERNAME_HASH = "C62E2BC30DC0FD097E2407D90DEE7445"  # ytbsydh 的 MD5
PASSWORD_HASH = "A1EAF07BB51B073BB923BD5BEB1AC8F8"  # 123456 的 MD5
def get_excel_path():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), f"kardz_orders_{date_str}.xlsx")

EXCEL_FILE = get_excel_path()
POLL_INTERVAL = 86400  # 轮询间隔（秒），每天一次
PAGE_SIZE    = 50     # 每页条数，越大请求次数越少
# ==============================

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Referer": "https://ch-admin.kardz.com/",
})

current_sessionid = ""  # 登录后保存，请求时带入 body

seen_ids = set()


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def login():
    """自动登录，成功返回 True"""
    global current_sessionid
    try:
        # 第一步：提交账号密码
        resp = session.post(
            f"{BASE_URL}/login",
            data={"username": USERNAME_HASH, "password": PASSWORD_HASH},
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        data = resp.json()
        if str(data.get("result")).lower() != "true":
            print(f"[{now()}] ❌ 登录失败: {data}")
            return False

        jsessionid = data.get("sessionid", "")
        current_sessionid = jsessionid
        session.cookies.set("JSESSIONID", jsessionid, domain="ch-admin.kardz.com")
        session.cookies.set("pageSize", str(PAGE_SIZE), domain="ch-admin.kardz.com")
        session.cookies.set("pageNo", "1", domain="ch-admin.kardz.com")
        print(f"[{now()}] ✅ 登录成功  JSESSIONID={jsessionid[:10]}...")

        # 第二步：请求 /index 初始化会话（服务器会在 Session 里设置 currentGroupId）
        time.sleep(0.5)
        session.get(f"{BASE_URL}/index", timeout=15)

        # 第三步：请求 /authInfo 进一步初始化（部分系统需要）
        time.sleep(0.3)
        session.get(f"{BASE_URL}/authInfo", timeout=10)

        print(f"[{now()}] ✅ 会话初始化完成")
        return True

    except Exception as e:
        print(f"[{now()}] ❌ 登录异常: {e}")
        return False


def is_session_expired(data):
    """判断是否 Session 过期"""
    if isinstance(data, dict):
        msg = str(data.get("message", "") + data.get("msg", "")).lower()
        if "login" in msg or "session" in msg or "expire" in msg or "未登录" in msg:
            return True
    return False


def fetch_page(page=1):
    """获取单页订单数据，返回 (list, total)，Session 过期返回 None"""
    try:
        end_time   = datetime.now().strftime("%Y-%m-%d 23:59:59")
        begin_time = "2020-01-01 00:00:00"   # 取全量历史
        body = (
            f"groupId=dada"
            f"&beginTime={begin_time}"
            f"&endTime={end_time}"
            f"&pageNo={page}"
            f"&pageSize={PAGE_SIZE}"
        )
        resp = session.post(
            f"{BASE_URL}/order/listData",
            data=body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        if "login" in resp.url:
            return None

        data = resp.json()
        if is_session_expired(data):
            return None

        if page == 1:
            print(f"[{now()}] 🔍 响应: count={data.get('count')}  list长度={len(data.get('list') or [])}")

        return data.get("list") or [], int(data.get("count") or 0)
    except Exception as e:
        print(f"[{now()}] ⚠️ 请求异常(页{page}): {e}")
        return None  # 触发重试


def fetch_all_orders():
    """翻页获取全部订单，返回列表；Session 过期返回 None"""
    result = fetch_page(1)
    if result is None:
        return None

    orders, total = result
    pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
    print(f"[{now()}] 共 {total} 条订单，共 {pages} 页")

    for page in range(2, pages + 1):
        result = None
        for retry in range(3):  # 最多重试3次
            result = fetch_page(page)
            if result is not None:
                break
            print(f"[{now()}] 🔄 第{page}页重试({retry+1}/3)...")
            time.sleep(2)
        if result is None:
            print(f"[{now()}] ❌ 第{page}页多次失败，跳过（下次轮询补全）")
            continue
        orders.extend(result[0])
        if page % 10 == 0:
            print(f"[{now()}] 📥 已抓取 {page}/{pages} 页，共 {len(orders)} 条...")
        time.sleep(0.3)

    return orders


def init_excel():
    """初始化 Excel，加载已有订单 ID"""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                seen_ids.add(row[0])
        print(f"[{now()}] 📂 已加载历史记录 {len(seen_ids)} 条")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "订单数据"

        # 表头样式
        headers = ["订单ID", "手机号", "邮箱", "商品名称", "金额(元)", "渠道", "下单时间", "记录时间"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 列宽
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20

        wb.save(EXCEL_FILE)
        print(f"[{now()}] 📄 创建新 Excel: {EXCEL_FILE}")


def append_new_orders(new_orders):
    """将新订单追加到 Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for o in new_orders:
        ws.append([
            o.get("id"),
            o.get("phoneNo", ""),
            o.get("emailAccount", ""),
            o.get("goodsName", ""),
            o.get("payAmount", ""),
            o.get("channel", ""),
            o.get("payTime", ""),
            now()
        ])
    wb.save(EXCEL_FILE)


def main():
    global EXCEL_FILE

    if not login():
        print("初始登录失败，请检查账号密码，退出。")
        return

    while True:
        try:
            # 每轮开始时按当天日期更新文件名
            EXCEL_FILE = get_excel_path()
            seen_ids.clear()
            print("=" * 50)
            print(f"  Kardz 订单监控  轮询间隔: {POLL_INTERVAL}秒")
            print(f"  输出文件: {EXCEL_FILE}")
            print("=" * 50)
            init_excel()

            orders = fetch_all_orders()

            if orders is None:
                print(f"[{now()}] 🔄 Session 过期，重新登录...")
                login()
                time.sleep(5)
                continue

            new_orders = [o for o in orders if o.get("id") not in seen_ids]

            if new_orders:
                append_new_orders(new_orders)
                for o in new_orders:
                    seen_ids.add(o.get("id"))
                phones = [o.get("phoneNo") for o in new_orders if o.get("phoneNo")]
                emails = [o.get("emailAccount") for o in new_orders if o.get("emailAccount")]
                print(f"[{now()}] 🆕 新增 {len(new_orders)} 条  手机: {phones}  邮箱: {emails}")
                print(f"[{now()}] 💾 Excel 已保存: {EXCEL_FILE}")
            else:
                print(f"[{now()}] ✔ 无新订单（总 {len(orders)} 条）")

        except KeyboardInterrupt:
            print(f"\n[{now()}] 手动停止。")
            break
        except Exception as e:
            print(f"[{now()}] ⚠️ 异常: {e}")

        time.sleep(POLL_INTERVAL)


if __name__ == "__main__":
    main()
