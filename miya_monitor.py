"""
Miya IP 邀请用户邮箱抓取脚本
- 自动登录，token 过期自动重登
- 全量抓取 + 每 5 分钟增量监控
- 新邮箱自动追加写入 Excel
"""

import requests
import time
import os
import platform
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============ 配置 ============
BASE_URL   = "https://www.miyaip.com"
EMAIL      = "terminos888@gmail.com"
PASSWORD   = "Qweasd123456.."
PAGE_SIZE  = 100
POLL_INTERVAL = 86400  # 每天一次

def get_excel_path():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), f"miya_emails_{date_str}.xlsx")

EXCEL_FILE = get_excel_path()
# ==============================

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "Origin": "https://www.miyaip.com",
    "Referer": "https://www.miyaip.com/",
})

current_token = ""
seen_ids = set()


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def login():
    global current_token
    try:
        resp = session.post(
            f"{BASE_URL}/api/SysLogin/Login",
            json={"email": EMAIL, "passWord": PASSWORD, "loginType": "pwd"},
            timeout=15
        )
        data = resp.json()

        # token 在 body.token 里
        token = (
            (data.get("body") or {}).get("token") or
            data.get("token") or
            (data.get("data") or {}).get("token") or
            ""
        )

        if token:
            current_token = token
            session.headers.update({"Authorization": f"Bearer {token}"})
            print(f"[{now()}] ✅ 登录成功  token={token[:16]}...")
            return True
        else:
            print(f"[{now()}] ❌ 登录失败，未找到 token，完整响应: {data}")
            return False
    except Exception as e:
        print(f"[{now()}] ❌ 登录异常: {e}")
        return False


def is_token_expired(data):
    if isinstance(data, dict):
        code = str(data.get("code", "") or data.get("Code", ""))
        msg  = str(data.get("msg", "") or data.get("message", "") or data.get("Message", "")).lower()
        if code in ("401", "403") or "token" in msg or "login" in msg or "expire" in msg or "未登录" in msg:
            return True
    return False


def fetch_page(page=1):
    try:
        resp = session.get(
            f"{BASE_URL}/api/InviteHistory/InvitedUserPageList",
            params={"PageNo": page, "PageSize": PAGE_SIZE},
            timeout=15
        )
        data = resp.json()

        if is_token_expired(data):
            return None

        body  = data.get("body") or {}
        lst   = body.get("records") or []
        total = body.get("totalRows") or 0

        return lst, int(total)
    except Exception as e:
        print(f"[{now()}] ⚠️ 请求异常(页{page}): {e}")
        return [], 0


def fetch_all():
    result = fetch_page(1)
    if result is None:
        return None

    items, total = result
    if total == 0 and not items:
        return items

    pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    print(f"[{now()}] 共 {total} 条，共 {pages} 页")

    for page in range(2, pages + 1):
        result = fetch_page(page)
        if result is None:
            return None
        items.extend(result[0])
        if page % 10 == 0:
            print(f"[{now()}] 📥 已抓取 {page}/{pages} 页，共 {len(items)} 条...")
        time.sleep(0.3)

    return items


def extract_fields(item):
    email   = item.get("email") or item.get("account") or ""
    account = item.get("account") or ""
    name    = item.get("name") or ""
    tel     = item.get("tel") or ""
    reg     = item.get("sysCreateTime") or ""
    return email, account, name, tel, reg


def init_excel():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                seen_ids.add(row[0])
        print(f"[{now()}] 📂 已加载历史记录 {len(seen_ids)} 条  文件: {EXCEL_FILE}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "邮箱数据"
        headers = ["邮箱", "账号", "用户名", "手机号", "注册时间", "记录时间"]
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        for col, w in zip("ABCDEF", [28, 28, 16, 14, 22, 22]):
            ws.column_dimensions[col].width = w
        wb.save(EXCEL_FILE)
        print(f"[{now()}] 📄 创建新 Excel: {EXCEL_FILE}")


def append_items(new_items):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for item in new_items:
        email, account, name, tel, reg = extract_fields(item)
        ws.append([email, account, name, tel, reg, now()])
    wb.save(EXCEL_FILE)


def main():
    global EXCEL_FILE

    if not login():
        print("初始登录失败，退出。")
        return

    while True:
        try:
            # 每轮开始时按当天日期更新文件名
            EXCEL_FILE = get_excel_path()
            seen_ids.clear()
            print("=" * 52)
            print("  Miya IP 邮箱监控")
            print(f"  输出文件: {EXCEL_FILE}")
            print("=" * 52)
            init_excel()

            items = fetch_all()

            if items is None:
                print(f"[{now()}] 🔄 Token 过期，重新登录...")
                login()
                time.sleep(3)
                continue

            new_items = [i for i in items if extract_fields(i)[0] not in seen_ids]

            if new_items:
                append_items(new_items)
                for i in new_items:
                    seen_ids.add(extract_fields(i)[0])
                emails = [extract_fields(i)[0] for i in new_items if extract_fields(i)[0]]
                print(f"[{now()}] 🆕 新增 {len(new_items)} 条  邮箱: {emails[:5]}{'...' if len(emails)>5 else ''}")
                print(f"[{now()}] 💾 已保存: {EXCEL_FILE}")
            else:
                print(f"[{now()}] ✔ 无新数据（总 {len(items)} 条）")

        except KeyboardInterrupt:
            print(f"\n[{now()}] 手动停止，共记录 {len(seen_ids)} 条。")
            break
        except Exception as e:
            print(f"[{now()}] ⚠️ 异常: {e}")

        time.sleep(POLL_INTERVAL)


if __name__ == "__main__":
    main()
